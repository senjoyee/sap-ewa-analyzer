"""Quick smoke-test for worker_synthesis_to_excel."""
import sys
import json
import io
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.excel_utils import worker_synthesis_to_excel
from openpyxl import load_workbook

synth_path = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "docs",
    "SHP_20474620_850648038_2026-03-02_R_EWA_worker_synthesis.json",
)
# resolve relative to repo root
synth_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "docs",
    "SHP_20474620_850648038_2026-03-02_R_EWA_worker_synthesis.json",
)

if not os.path.exists(synth_path):
    print(f"SKIP: artifact not found at {synth_path}")
    sys.exit(0)

with open(synth_path, encoding="utf-8") as f:
    data = json.load(f)

b = worker_synthesis_to_excel(data, customer_name="TestCustomer")
print(f"Workbook size: {len(b)} bytes")

wb = load_workbook(io.BytesIO(b))
print("Sheets:", wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s}: {ws.max_row} rows x {ws.max_column} cols")

# Basic assertions
assert "Summary" in wb.sheetnames, "Missing Summary sheet"
assert "Traceability" in wb.sheetnames, "Missing Traceability sheet"
assert "Open Questions" in wb.sheetnames, "Missing Open Questions sheet"
assert len(wb.sheetnames) >= 4, f"Expected >=4 sheets, got {len(wb.sheetnames)}"

summary_ws = wb["Summary"]
assert summary_ws["A1"].value == "EWA Worker Synthesis Report"

traceability_ws = wb["Traceability"]
assert traceability_ws.max_row > 4, "Traceability sheet has no data rows"

print("\nAll assertions passed.")
