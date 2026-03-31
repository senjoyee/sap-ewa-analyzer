"""Workbook builder for the V2 agentic pipeline.

Produces: Executive Summary + up to 6 domain tabs (Security, Database, Performance,
Basis, Business, Lifecycle). The Business tab is omitted when the system is not
ECC or S/4HANA. Each domain tab has three sections: Findings, Parameters, and
AI Deep Analysis supplements.
"""
from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agent.specialist_agents import DomainResult
from agent.deep_thinker_agent import SupplementalFinding
from utils.ewa_slicer import ChapterData

logger = logging.getLogger(__name__)

# SAP-inspired color palette (shared with excel_utils.py)
COLORS = {
    "sap_gold": "F0AB00",
    "sap_dark": "333333",
    "header_bg": "E5E5E5",
    "white": "FFFFFF",
    "light_gray": "F9F9F9",
    "border": "CCCCCC",
    "red": "CC0000",
    "yellow": "E07000",
    "implicit": "6868AC",
    "green": "008000",
    "section_bg": "D6DCE4",
}

DOMAIN_DISPLAY_NAMES: dict[str, str] = {
    "security": "Security",
    "database": "Database",
    "performance": "Performance",
    "basis": "Basis",
    "business": "Business",
    "lifecycle": "Lifecycle",
}

DOMAIN_TAB_ORDER = ["security", "database", "performance", "basis", "business", "lifecycle"]

# Shared styling helpers
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="333333")
_HEADER_FILL = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
_HEADER_BORDER = Border(
    bottom=Side(style="medium", color="333333"),
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)
_DATA_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)
_WRAP_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _is_business_applicable(dr: DomainResult) -> bool:
    """Return False when the business specialist signalled the system is not ECC/S4."""
    if not dr.applicable:
        return False
    if any(a.get("reason") == "not_applicable_system_type" for a in dr.abstentions):
        return False
    return True


def build_workbook(
    domain_results: list[DomainResult],
    supplemental_findings: list[SupplementalFinding],
    domain_chapters: dict[str, list[ChapterData]] | None = None,
    system_metadata: dict[str, str] | None = None,
) -> bytes:
    """Build a workbook from pipeline outputs.

    The Business tab is omitted entirely when the system is not ECC/S4.
    Returns xlsx bytes ready for blob storage.
    """
    wb = Workbook()
    metadata = system_metadata or {}
    chapters_by_domain = domain_chapters or {}

    # Index results by domain
    results_by_domain: dict[str, DomainResult] = {}
    for dr in domain_results:
        results_by_domain[dr.domain] = dr

    # Determine whether business tab should be included
    business_dr = results_by_domain.get("business", DomainResult(domain="business"))
    include_business = _is_business_applicable(business_dr)

    # Index supplemental findings by domain
    supplements_by_domain: dict[str, list[dict[str, Any]]] = {d: [] for d in DOMAIN_TAB_ORDER}
    for sf in supplemental_findings:
        d = sf.domain if isinstance(sf, SupplementalFinding) else sf.get("domain", "")
        sd = sf.to_dict() if isinstance(sf, SupplementalFinding) else sf
        if d in supplements_by_domain:
            supplements_by_domain[d].append(sd)

    # Tab 1: Executive Summary
    ws_exec = wb.active
    _write_executive_summary(ws_exec, results_by_domain, supplemental_findings, metadata, include_business)

    # Remaining domain tabs (business skipped when not applicable)
    for domain in DOMAIN_TAB_ORDER:
        if domain == "business":
            if not include_business:
                continue
            ws = wb.create_sheet(title=DOMAIN_DISPLAY_NAMES[domain])
            sups = supplements_by_domain.get(domain, [])
            _write_business_tab(ws, business_dr, sups, chapters_by_domain.get(domain, []))
        elif domain == "performance":
            ws = wb.create_sheet(title=DOMAIN_DISPLAY_NAMES[domain])
            dr = results_by_domain.get(domain, DomainResult(domain=domain))
            sups = supplements_by_domain.get(domain, [])
            _write_performance_tab(ws, dr, sups, chapters_by_domain.get(domain, []))
        else:
            ws = wb.create_sheet(title=DOMAIN_DISPLAY_NAMES[domain])
            dr = results_by_domain.get(domain, DomainResult(domain=domain))
            sups = supplements_by_domain.get(domain, [])
            _write_domain_tab(ws, dr, sups, domain, chapters_by_domain.get(domain, []))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Executive Summary Tab
# ---------------------------------------------------------------------------

def _write_executive_summary(
    ws: Worksheet,
    results_by_domain: dict[str, DomainResult],
    supplemental_findings: list[SupplementalFinding],
    metadata: dict[str, str],
    include_business: bool = True,
):
    ws.title = "Executive Summary"

    # Title
    ws["A1"] = "EWA Workbook — Executive Summary"
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30

    # System Metadata
    row = 3
    ws[f"A{row}"] = "System Information"
    ws[f"A{row}"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    for label, key in [("System ID", "system_id"), ("Report Date", "report_date"), ("Analysis Period", "analysis_period")]:
        val = metadata.get(key, metadata.get(label, ""))
        if val:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color=COLORS["sap_dark"])
            ws[f"B{row}"] = str(val)
            row += 1
    row += 1

    # Domain Overview Matrix
    ws[f"A{row}"] = "Domain Overview"
    ws[f"A{row}"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    matrix_headers = ["Domain", "Findings / Observations", "Parameters", "Derived Recommendations"]
    for col_idx, header in enumerate(matrix_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        _apply_header(cell)
    row += 1

    for domain in DOMAIN_TAB_ORDER:
        dr = results_by_domain.get(domain, DomainResult(domain=domain))

        if domain == "business" and not include_business:
            values = [DOMAIN_DISPLAY_NAMES.get(domain, domain), "N/A — not applicable for this system type", "", ""]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
                if col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                if col_idx == 2:
                    cell.font = Font(name="Calibri", size=11, italic=True, color="888888")
            row += 1
            continue

        finding_count = len(dr.findings)
        param_count = len(dr.parameters)
        sup_count = sum(1 for sf in supplemental_findings
                        if (sf.domain if isinstance(sf, SupplementalFinding) else sf.get("domain", "")) == domain)

        values = [DOMAIN_DISPLAY_NAMES.get(domain, domain), finding_count, param_count, sup_count]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            _apply_data(cell, row)
            if col_idx == 1:
                cell.font = Font(name="Calibri", size=11, bold=True)
        row += 1
    row += 1

    # Priority Recommendations (non-business domains only — business has its own status report tab)
    ws[f"A{row}"] = "Priority Recommendations"
    ws[f"A{row}"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    # Collect actionable findings across non-business domains
    all_findings: list[tuple[str, dict]] = []
    for domain in DOMAIN_TAB_ORDER:
        if domain == "business":
            continue
        dr = results_by_domain.get(domain, DomainResult(domain=domain))
        for f in dr.findings:
            all_findings.append((domain, f))

    risk_headers = ["ID", "Domain", "Title", "Finding", "Impact", "Recommendation"]
    for col_idx, header in enumerate(risk_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        _apply_header(cell)
    row += 1

    for domain, finding in all_findings[:5]:
        vals = [
            finding.get("finding_id", ""),
            DOMAIN_DISPLAY_NAMES.get(domain, domain),
            finding.get("title", ""),
            finding.get("finding", ""),
            finding.get("impact", ""),
            finding.get("recommendation", ""),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
            _apply_data(cell, row)
        row += 1

    # Deep Thinker Highlights
    if supplemental_findings:
        row += 1
        ws[f"A{row}"] = "AI Deep Analysis Highlights"
        ws[f"A{row}"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        dt_headers = ["ID", "Domain", "Title", "Finding", "Rationale", "Recommendation"]
        for col_idx, header in enumerate(dt_headers, start=1):
            if header:
                cell = ws.cell(row=row, column=col_idx, value=header)
                _apply_header(cell)
        row += 1

        for sf in supplemental_findings[:3]:
            sd = sf.to_dict() if isinstance(sf, SupplementalFinding) else sf
            vals = [
                sd.get("finding_id", ""),
                DOMAIN_DISPLAY_NAMES.get(sd.get("domain", ""), sd.get("domain", "")),
                sd.get("title", ""),
                sd.get("finding", ""),
                sd.get("rationale", ""),
                sd.get("recommendation", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                _apply_data(cell, row)
            row += 1

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Domain Tab
# ---------------------------------------------------------------------------

def _write_domain_tab(
    ws: Worksheet,
    dr: DomainResult,
    supplements: list[dict[str, Any]],
    domain: str,
    chapters: list[ChapterData],
):
    display_name = DOMAIN_DISPLAY_NAMES.get(domain, domain)
    row = 1

    # Title
    ws[f"A{row}"] = f"{display_name} — Domain Analysis"
    ws[f"A{row}"].font = Font(name="Calibri", size=18, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 28
    row += 2

    # Section A: Findings
    row = _write_section_header(ws, row, "A — Findings")
    if dr.findings:
        headers = ["ID", "Source Chapter", "Title", "Finding", "Impact", "Recommendation"]
        row = _write_table_headers(ws, row, headers)
        for finding in dr.findings:
            vals = [
                finding.get("finding_id", ""),
                finding.get("source_chapter", ""),
                finding.get("title", ""),
                finding.get("finding", ""),
                finding.get("impact", ""),
                finding.get("recommendation", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                _apply_data(cell, row)
            ws.row_dimensions[row].height = 60
            row += 1
    else:
        ws[f"A{row}"] = "No actionable findings in this domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1
    row += 1

    # Section B: Parameters
    row = _write_section_header(ws, row, "B — Parameters")
    if dr.parameters:
        headers = ["Parameter", "Current Value", "Recommended", "Action", "Source Chapter", ""]
        row = _write_table_headers(ws, row, headers)
        for param in dr.parameters:
            vals = [
                param.get("param_name", ""),
                param.get("current_value", ""),
                param.get("recommended_value", ""),
                param.get("action", ""),
                param.get("source_chapter", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
            row += 1
    else:
        ws[f"A{row}"] = "No parameter changes recommended in this domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1
    row += 1

    # Section C: Deep Thinker Supplements
    row = _write_section_header(ws, row, "C — AI Deep Analysis (Derived Recommendations)")
    if supplements:
        headers = ["ID", "Title", "Finding", "Rationale", "Recommendation", ""]
        row = _write_table_headers(ws, row, headers)
        for sf in supplements:
            vals = [
                sf.get("finding_id", ""),
                sf.get("title", ""),
                sf.get("finding", ""),
                sf.get("rationale", ""),
                sf.get("recommendation", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                _apply_data(cell, row)
                if col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, color=COLORS["implicit"])
            ws.row_dimensions[row].height = 60
            row += 1
    else:
        ws[f"A{row}"] = "No derived recommendations for this domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1

    row += 1

    # Section D: Chapters Considered
    row = _write_section_header(ws, row, "D — Chapters Considered")
    if chapters:
        headers = ["Chapter", "Chapter Name", "", "", "", ""]
        row = _write_table_headers(ws, row, headers)
        for chapter in chapters:
            vals = [chapter.number, chapter.title]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
            row += 1
    else:
        ws[f"A{row}"] = "No chapters were routed to this domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Business Status Report Tab (reporting-first, separate layout)
# ---------------------------------------------------------------------------

_RAG_FILL: dict[str, str] = {
    "RED": COLORS["red"],
    "YELLOW": COLORS["yellow"],
    "GREEN": COLORS["green"],
}
_RAG_TEXT_COLOR: dict[str, str] = {
    "RED": COLORS["white"],
    "YELLOW": COLORS["white"],
    "GREEN": COLORS["white"],
}


def _apply_rag_status_cell(cell, rag_status: str | None):
    """Apply RAG traffic-light fill to a status cell."""
    val = (rag_status or "").upper()
    fill_color = _RAG_FILL.get(val)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color=_RAG_TEXT_COLOR.get(val, COLORS["white"]))
    else:
        # Unknown / null — neutral styling
        cell.font = Font(name="Calibri", size=11, italic=True, color="888888")
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = _DATA_BORDER
    cell.value = val if fill_color else (rag_status or "—")


def _write_business_tab(
    ws: Worksheet,
    dr: DomainResult,
    supplements: list[dict[str, Any]],
    chapters: list[ChapterData],
):
    row = 1

    # Title
    ws[f"A{row}"] = "Business — Status Report"
    ws[f"A{row}"].font = Font(name="Calibri", size=18, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 28
    row += 2

    # Section A: Business Observations
    row = _write_section_header(ws, row, "A — Business Area Status")
    if dr.findings:
        headers = ["ID", "Status", "Topic Area", "Observation", "Business Significance", "Action Required"]
        row = _write_table_headers(ws, row, headers)
        for obs in dr.findings:
            rag = obs.get("rag_status")
            recommendation = obs.get("recommendation") or ""
            row_vals = [
                obs.get("finding_id", ""),
                None,  # placeholder — written by _apply_rag_status_cell
                obs.get("title", ""),
                obs.get("finding", ""),
                obs.get("impact", ""),
                recommendation if recommendation and recommendation.lower() not in ("null", "none") else "No action required",
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                if col_idx == 2:
                    cell = ws.cell(row=row, column=col_idx)
                    _apply_rag_status_cell(cell, rag)
                else:
                    cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                    _apply_data(cell, row)
            ws.row_dimensions[row].height = 60
            row += 1
    else:
        ws[f"A{row}"] = "No business observations available."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1
    row += 1

    # Section B: Parameters
    row = _write_section_header(ws, row, "B — Parameters")
    if dr.parameters:
        headers = ["Parameter", "Current Value", "Recommended", "Action", "Source Chapter", ""]
        row = _write_table_headers(ws, row, headers)
        for param in dr.parameters:
            vals = [
                param.get("param_name", ""),
                param.get("current_value", ""),
                param.get("recommended_value", ""),
                param.get("action", ""),
                param.get("source_chapter", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
            row += 1
    else:
        ws[f"A{row}"] = "No parameter changes recommended in the business domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1
    row += 1

    # Section C: Deep Thinker Supplements
    row = _write_section_header(ws, row, "C — AI Deep Analysis (Derived Recommendations)")
    if supplements:
        headers = ["ID", "Title", "Finding", "Rationale", "Recommendation", ""]
        row = _write_table_headers(ws, row, headers)
        for sf in supplements:
            vals = [
                sf.get("finding_id", ""),
                sf.get("title", ""),
                sf.get("finding", ""),
                sf.get("rationale", ""),
                sf.get("recommendation", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                _apply_data(cell, row)
                if col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, color=COLORS["implicit"])
            ws.row_dimensions[row].height = 60
            row += 1
    else:
        ws[f"A{row}"] = "No derived recommendations for the business domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1

    row += 1

    # Section D: Chapters Considered
    row = _write_section_header(ws, row, "D — Chapters Considered")
    if chapters:
        headers = ["Chapter", "Chapter Name", "", "", "", ""]
        row = _write_table_headers(ws, row, headers)
        for chapter in chapters:
            vals = [chapter.number, chapter.title]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
            row += 1
    else:
        ws[f"A{row}"] = "No chapters were routed to the business domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Performance Status Report Tab (observation-first, with RAG status)
# ---------------------------------------------------------------------------

def _write_performance_tab(
    ws: Worksheet,
    dr: DomainResult,
    supplements: list[dict[str, Any]],
    chapters: list[ChapterData],
):
    row = 1

    # Title
    ws[f"A{row}"] = "Performance — System Status Report"
    ws[f"A{row}"].font = Font(name="Calibri", size=18, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 28
    row += 2

    # Section A: Performance Observations
    row = _write_section_header(ws, row, "A — Performance Status")
    if dr.findings:
        headers = ["ID", "Status", "Topic Area", "Observation", "Significance", "Action Required"]
        row = _write_table_headers(ws, row, headers)
        for obs in dr.findings:
            rag = obs.get("rag_status")
            recommendation = obs.get("recommendation") or ""
            row_vals = [
                obs.get("finding_id", ""),
                None,  # placeholder — written by _apply_rag_status_cell
                obs.get("title", ""),
                obs.get("finding", ""),
                obs.get("impact", ""),
                recommendation if recommendation and recommendation.lower() not in ("null", "none") else "No action required",
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                if col_idx == 2:
                    cell = ws.cell(row=row, column=col_idx)
                    _apply_rag_status_cell(cell, rag)
                else:
                    cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                    _apply_data(cell, row)
            ws.row_dimensions[row].height = 60
            row += 1
    else:
        ws[f"A{row}"] = "No performance observations available."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1
    row += 1

    # Section B: Parameters
    row = _write_section_header(ws, row, "B — Parameters")
    if dr.parameters:
        headers = ["Parameter", "Current Value", "Recommended", "Action", "Source Chapter", ""]
        row = _write_table_headers(ws, row, headers)
        for param in dr.parameters:
            vals = [
                param.get("param_name", ""),
                param.get("current_value", ""),
                param.get("recommended_value", ""),
                param.get("action", ""),
                param.get("source_chapter", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
            row += 1
    else:
        ws[f"A{row}"] = "No parameter changes recommended in the performance domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1
    row += 1

    # Section C: Deep Thinker Supplements
    row = _write_section_header(ws, row, "C — AI Deep Analysis (Derived Recommendations)")
    if supplements:
        headers = ["ID", "Title", "Finding", "Rationale", "Recommendation", ""]
        row = _write_table_headers(ws, row, headers)
        for sf in supplements:
            vals = [
                sf.get("finding_id", ""),
                sf.get("title", ""),
                sf.get("finding", ""),
                sf.get("rationale", ""),
                sf.get("recommendation", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                _apply_data(cell, row)
                if col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, color=COLORS["implicit"])
            ws.row_dimensions[row].height = 60
            row += 1
    else:
        ws[f"A{row}"] = "No derived recommendations for the performance domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1

    row += 1

    # Section D: Chapters Considered
    row = _write_section_header(ws, row, "D — Chapters Considered")
    if chapters:
        headers = ["Chapter", "Chapter Name", "", "", "", ""]
        row = _write_table_headers(ws, row, headers)
        for chapter in chapters:
            vals = [chapter.number, chapter.title]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _apply_data(cell, row)
            row += 1
    else:
        ws[f"A{row}"] = "No chapters were routed to the performance domain."
        ws[f"A{row}"].font = Font(name="Calibri", size=11, italic=True, color="666666")
        row += 1

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Shared cell formatting helpers
# ---------------------------------------------------------------------------

def _apply_header(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _HEADER_BORDER
    cell.alignment = _WRAP_ALIGNMENT


def _apply_data(cell, row_num: int):
    cell.font = Font(name="Calibri", size=11)
    cell.border = _DATA_BORDER
    cell.alignment = _WRAP_ALIGNMENT
    if row_num % 2 == 0:
        cell.fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")


def _write_section_header(ws: Worksheet, row: int, text: str) -> int:
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = Font(name="Calibri", size=13, bold=True, color=COLORS["sap_dark"])
    ws[f"A{row}"].fill = PatternFill(start_color=COLORS["section_bg"], end_color=COLORS["section_bg"], fill_type="solid")
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 24
    return row + 1


def _write_table_headers(ws: Worksheet, row: int, headers: list[str]) -> int:
    for col_idx, header in enumerate(headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            _apply_header(cell)
    return row + 1


def _auto_fit(ws: Worksheet, min_width: int = 14, max_width: int = 55):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                val = str(cell.value) if cell.value else ""
                lines = val.split("\n")
                max_line = max(len(line) for line in lines) if lines else 0
                max_length = max(max_length, max_line)
            except Exception:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted
