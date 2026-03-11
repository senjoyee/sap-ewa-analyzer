"""
Excel export utilities for EWA analysis data.
Generates professionally formatted Excel workbooks from EWA JSON.
"""

from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation


# SAP-inspired color palette
COLORS = {
    "sap_gold": "F0AB00",
    "sap_dark": "333333",
    "header_bg": "E5E5E5",
    "white": "FFFFFF",
    "light_gray": "F9F9F9",
    "border": "CCCCCC",
    "critical": "CC0000",
    "high": "E07000",
    "medium": "F0AB00",
    "low": "008000",
    "good": "008000",
    "fair": "E07000",
    "poor": "CC0000",
}

CONFIG_NAME_HINTS = (
    ".ini",
    "[",
    "]",
    "/",
    "profile",
    "parameter",
    "limit",
    "timeout",
    "buffer",
    "ssl",
    "tls",
    "listeninterface",
    "hostname_resolution",
    "allocation",
    "memory",
)

ACTION_NAME_HINTS = (
    "privilege",
    "release",
    "patch",
    "kernel",
    "monitor",
    "backup",
    "recovery",
    "restore",
    "assignment",
    "upgrade",
    "housekeeping",
    "baseline",
    "runbook",
    "role",
    "authorization",
    "user",
)

ACTION_TEXT_HINTS = (
    "review",
    "ensure",
    "perform",
    "restrict",
    "assign",
    "remove",
    "update",
    "upgrade",
    "monitor",
    "document",
    "validate",
    "align",
    "establish",
    "restart",
    "re-validate",
    "cleanup",
    "close",
)


def _create_styles() -> Dict[str, NamedStyle]:
    """Create reusable named styles for the workbook."""
    styles = {}
    
    # Title style (for sheet titles)
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_dark"])
    title_style.alignment = Alignment(horizontal="left", vertical="center")
    styles["title"] = title_style
    
    # Header style (for table headers)
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name="Calibri", size=11, bold=True, color=COLORS["sap_dark"])
    header_style.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_style.border = Border(
        bottom=Side(style="medium", color=COLORS["sap_dark"]),
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
    )
    header_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    styles["header"] = header_style
    
    # Data cell style
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(name="Calibri", size=11)
    data_style.border = Border(
        bottom=Side(style="thin", color=COLORS["border"]),
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
    )
    data_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    styles["data"] = data_style
    
    # Label style (for key-value pairs)
    label_style = NamedStyle(name="label_style")
    label_style.font = Font(name="Calibri", size=11, bold=True, color=COLORS["sap_dark"])
    label_style.alignment = Alignment(horizontal="left", vertical="top")
    styles["label"] = label_style
    
    # Value style
    value_style = NamedStyle(name="value_style")
    value_style.font = Font(name="Calibri", size=11)
    value_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    styles["value"] = value_style
    
    return styles


def _apply_header_style(cell, styles: Dict[str, NamedStyle]):
    """Apply header styling to a cell."""
    style = styles["header"]
    cell.font = style.font
    cell.fill = style.fill
    cell.border = style.border
    cell.alignment = style.alignment


def _apply_data_style(cell, styles: Dict[str, NamedStyle], alt_row: bool = False):
    """Apply data cell styling."""
    style = styles["data"]
    cell.font = style.font
    cell.border = style.border
    cell.alignment = style.alignment
    if alt_row:
        cell.fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")


def _get_severity_fill(severity: str) -> PatternFill:
    """Get fill color based on severity level."""
    severity_lower = severity.lower() if severity else ""
    color = COLORS.get(severity_lower, COLORS["medium"])
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _get_status_fill(status: str) -> PatternFill:
    """Get fill color based on status (good/fair/poor)."""
    status_lower = status.lower() if status else ""
    color = COLORS.get(status_lower, COLORS["sap_gold"])
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _auto_fit_columns(ws: Worksheet, min_width: int = 12, max_width: int = 60):
    """Auto-fit column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value else ""
                # Account for line breaks
                lines = cell_value.split("\n")
                max_line_length = max(len(line) for line in lines) if lines else 0
                max_length = max(max_length, max_line_length)
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column].width = adjusted_width


def _write_summary_sheet(ws: Worksheet, data: Dict[str, Any], styles: Dict[str, NamedStyle], customer_name: str = ""):
    """Write the Summary sheet with metadata and executive summary."""
    ws.title = "Summary"
    
    # Title
    ws["A1"] = "EWA Analysis Summary"
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    
    # System Metadata section
    ws["A3"] = "System Information"
    ws["A3"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells("A3:D3")
    
    meta = data.get("System Metadata", data.get("system_metadata", {})) or {}
    row = 4
    
    # Add customer name if provided
    if customer_name:
        ws[f"A{row}"] = "Customer"
        ws[f"A{row}"].font = styles["label"].font
        ws[f"B{row}"] = customer_name
        ws[f"B{row}"].font = styles["value"].font
        row += 1
    
    meta_fields = [
        ("System ID", meta.get("System ID", meta.get("system_id", "N/A"))),
        ("Report Date", meta.get("Report Date", meta.get("report_date", "N/A"))),
        ("Analysis Period", meta.get("Analysis Period", meta.get("analysis_period", "N/A"))),
    ]
    
    for label, value in meta_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = styles["label"].font
        ws[f"B{row}"] = str(value)
        ws[f"B{row}"].font = styles["value"].font
        row += 1
    
    # Overall Risk
    row += 1
    ws[f"A{row}"] = "Overall Risk"
    ws[f"A{row}"].font = styles["label"].font
    overall_risk = data.get("Overall Risk", data.get("overall_risk", "N/A"))
    ws[f"B{row}"] = str(overall_risk).upper()
    ws[f"B{row}"].font = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
    ws[f"B{row}"].fill = _get_severity_fill(str(overall_risk))
    ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    # System Health Overview
    ws[f"A{row}"] = "System Health Overview"
    ws[f"A{row}"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    
    health = data.get("System Health Overview", data.get("system_health_overview", {})) or {}
    # Deduplicate configuration entry
    health_areas = ["Performance", "Security", "Stability", "Configuration"]
    
    for area in health_areas:
        status = health.get(area, health.get(area.lower()))
        if status:
            ws[f"A{row}"] = area.title()
            ws[f"A{row}"].font = styles["label"].font
            ws[f"B{row}"] = str(status).upper()
            ws[f"B{row}"].font = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
            ws[f"B{row}"].fill = _get_status_fill(str(status))
            ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
            row += 1
    
    row += 1
    
    # Executive Summary
    ws[f"A{row}"] = "Executive Summary"
    ws[f"A{row}"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    
    exec_summary = data.get("Executive Summary", data.get("executive_summary", ""))
    if exec_summary:
        ws[f"A{row}"] = exec_summary
        ws[f"A{row}"].font = styles["value"].font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = max(60, len(exec_summary) // 3)
    
    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _write_positive_findings_sheet(ws: Worksheet, data: Dict[str, Any], styles: Dict[str, NamedStyle]):
    """Write the Positive Findings sheet."""
    ws.title = "Positive Findings"
    
    findings = data.get("Positive Findings", data.get("positive_findings", [])) or []
    
    # Title
    ws["A1"] = "Positive Findings"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 25
    
    if not findings:
        ws["A3"] = "No positive findings recorded."
        return
    
    summary_groups = _group_positive_findings(findings)
    current_row = 3

    if summary_groups:
        ws[f"A{current_row}"] = "Summary by Area"
        ws[f"A{current_row}"].font = Font(name="Calibri", size=12, bold=True, color=COLORS["sap_dark"])
        ws.merge_cells(f"A{current_row}:B{current_row}")
        current_row += 1

        for area, descriptions in summary_groups:
            summary_text = "\n".join(f"- {desc}" for desc in descriptions)
            cell_a = ws.cell(row=current_row, column=1, value=area)
            cell_b = ws.cell(row=current_row, column=2, value=summary_text)

            cell_a.font = styles["label"].font
            _apply_data_style(cell_b, styles)
            cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            current_row += 1
    else:
        ws[f"A{current_row}"] = "No positive findings recorded."
        ws[f"A{current_row}"].font = styles["value"].font
        ws.merge_cells(f"A{current_row}:B{current_row}")
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


def _group_positive_findings(findings: List[Dict[str, Any]]) -> List[tuple[str, List[str]]]:
    """Group positive findings by area while preserving order."""
    groups: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for item in findings:
        area_raw = item.get("Area", item.get("area", "")) or "General"
        area = str(area_raw).strip() or "General"
        key = area.lower()
        if key not in groups:
            groups[key] = {"label": area, "items": [], "seen": set()}
            order.append(key)

        desc_raw = item.get("Description", item.get("description", "")) or ""
        desc = str(desc_raw).strip()
        if desc:
            normalized = re.sub(r"\s+", " ", desc).strip().lower()
            if normalized and normalized not in groups[key]["seen"]:
                groups[key]["items"].append(desc)
                groups[key]["seen"].add(normalized)

    return [(groups[k]["label"], groups[k]["items"]) for k in order]


def _write_findings_and_recommendations_sheet(ws: Worksheet, data: Dict[str, Any], styles: Dict[str, NamedStyle]):
    """Write combined Findings & Recommendations sheet grouped by severity."""
    ws.title = "Findings & Actions"
    
    findings = data.get("Key Findings", data.get("key_findings", [])) or []
    recs = data.get("Recommendations", data.get("recommendations", [])) or []
    
    # Build lookup of recommendations by linked issue ID
    recs_by_issue = {}
    for rec in recs:
        linked_id = rec.get("Linked issue ID", rec.get("linked_issue_id", ""))
        if linked_id:
            if linked_id not in recs_by_issue:
                recs_by_issue[linked_id] = []
            recs_by_issue[linked_id].append(rec)
    
    # Title
    ws["A1"] = "Findings & Recommended Actions"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 25
    
    if not findings:
        ws["A3"] = "No key findings recorded."
        return
    
    # Group findings by severity
    severity_order = ["critical", "high", "medium", "low"]
    grouped = {sev: [] for sev in severity_order}
    for finding in findings:
        sev = str(finding.get("Severity", finding.get("severity", "medium"))).lower()
        if sev not in grouped:
            sev = "medium"
        grouped[sev].append(finding)
    
    # Combined headers
    headers = [
        "Issue ID", "Area", "Finding", "Impact", "Business Impact", "Source",
        "Action", "Preventative Action", "Effort"
    ]
    
    current_row = 3
    group_ranges = {}
    
    for severity in severity_order:
        group_findings = grouped[severity]
        if not group_findings:
            continue
        
        # Severity group header row
        ws.cell(row=current_row, column=1, value=f"{severity.upper()} ({len(group_findings)})")
        ws.merge_cells(f"A{current_row}:I{current_row}")
        header_cell = ws[f"A{current_row}"]
        header_cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
        header_cell.fill = _get_severity_fill(severity)
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Column headers for this group
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            _apply_header_style(cell, styles)
        current_row += 1
        
        # Data rows for this severity
        data_start_row = current_row
        for idx, finding in enumerate(group_findings):
            issue_id = finding.get("Issue ID", finding.get("issue_id", ""))
            
            # Get linked recommendations
            linked_recs = recs_by_issue.get(issue_id, [])
            if linked_recs:
                rec = linked_recs[0]  # Take first linked recommendation
                effort = rec.get("Estimated Effort", rec.get("estimated_effort", {})) or {}
                if isinstance(effort, str):
                    effort_str = effort
                else:
                    effort_str = f"Analysis: {effort.get('analysis', 'N/A')}, Impl: {effort.get('implementation', 'N/A')}"
                action = rec.get("Action", rec.get("action", ""))
                preventative = rec.get("Preventative Action", rec.get("preventative_action", ""))
            else:
                action = ""
                preventative = ""
                effort_str = ""
            
            values = [
                issue_id,
                finding.get("Area", finding.get("area", "")),
                finding.get("Finding", finding.get("finding", "")),
                finding.get("Impact", finding.get("impact", "")),
                finding.get("Business impact", finding.get("business_impact", "")),
                finding.get("Source", finding.get("source", "")),
                action,
                preventative,
                effort_str,
            ]
            
            alt_row = idx % 2 == 1
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=str(value) if value else "")
                _apply_data_style(cell, styles, alt_row)
            
            ws.row_dimensions[current_row].height = 70
            current_row += 1
        
        data_end_row = current_row - 1
        group_ranges[severity] = (data_start_row, data_end_row)
        
        # Add blank row between groups
        current_row += 1
    
    # Apply Excel outline grouping - Critical expanded, others collapsed
    ws.sheet_properties.outlinePr.summaryBelow = False
    for severity, (start, end) in group_ranges.items():
        if start <= end:
            ws.row_dimensions.group(start, end, outline_level=1, hidden=(severity != "critical"))
    
    # Column widths
    widths = [10, 18, 40, 30, 30, 25, 40, 40, 20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_capacity_sheet(ws: Worksheet, data: Dict[str, Any], styles: Dict[str, NamedStyle]):
    """Write the Capacity Outlook sheet."""
    ws.title = "Capacity Outlook"
    
    capacity = data.get("Capacity Outlook", data.get("capacity_outlook", {})) or {}
    
    # Title
    ws["A1"] = "Capacity Outlook"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 25
    
    if not capacity:
        ws["A3"] = "No capacity data available."
        return
    
    # Key-value pairs
    fields = [
        ("Database Growth", capacity.get("Database Growth", capacity.get("database_growth", "N/A"))),
        ("CPU Utilization", capacity.get("CPU Utilization", capacity.get("cpu_utilization", "N/A"))),
        ("Memory Utilization", capacity.get("Memory Utilization", capacity.get("memory_utilization", "N/A"))),
    ]
    
    row = 3
    for label, value in fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = styles["label"].font
        ws[f"B{row}"] = str(value) if value else "N/A"
        ws[f"B{row}"].font = styles["value"].font
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Summary
    row += 1
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True, color=COLORS["sap_dark"])
    row += 1
    
    summary = capacity.get("Summary", capacity.get("summary", ""))
    if summary:
        ws[f"A{row}"] = summary
        ws[f"A{row}"].font = styles["value"].font
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = max(40, len(summary) // 2)
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


def _write_chapters_sheet(ws: Worksheet, data: Dict[str, Any], styles: Dict[str, NamedStyle]):
    """Write the Chapters Reviewed sheet."""
    ws.title = "Chapters Reviewed"
    
    chapters = data.get("Chapters Reviewed", data.get("chapters_reviewed", [])) or []
    
    # Title
    ws["A1"] = "Chapters Reviewed"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.row_dimensions[1].height = 25
    
    if not chapters:
        ws["A3"] = "No chapters information available."
        return
    
    # List chapters (no bullet prefix)
    for row_idx, chapter in enumerate(chapters, 3):
        cell = ws.cell(row=row_idx, column=1, value=f"{chapter}")
        cell.font = styles["value"].font
    
    ws.column_dimensions["A"].width = 60


def _write_parameters_sheet(ws: Worksheet, parameters: List[Dict[str, Any]], styles: Dict[str, NamedStyle]):
    """Write the Parameter Changes sheet with all recommended parameter modifications."""
    ws.title = "Parameter Changes"
    
    # Title
    ws["A1"] = "Recommended Parameter Changes"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 25
    
    if not parameters:
        ws["A3"] = "No parameter recommendations found in the report."
        return
    
    # Group parameters by area
    area_order = [
        "SAP HANA", "Database", "SAP Kernel", "Profile Parameters",
        "Application", "Memory/Buffer", "Operating System", "Network", "General"
    ]
    
    grouped = {}
    for param in parameters:
        area = param.get("area", "General")
        if area not in grouped:
            grouped[area] = []
        grouped[area].append(param)
    
    # Headers
    headers = ["Parameter Name", "Area", "Current Value", "Recommended Value", "Description", "Section"]
    
    current_row = 3
    
    # Write parameters grouped by area
    for area in area_order:
        if area not in grouped:
            continue
        
        area_params = grouped[area]
        
        # Area header row
        ws.cell(row=current_row, column=1, value=f"{area} ({len(area_params)} parameters)")
        ws.merge_cells(f"A{current_row}:F{current_row}")
        area_cell = ws[f"A{current_row}"]
        area_cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
        area_cell.fill = PatternFill(start_color=COLORS["sap_dark"], end_color=COLORS["sap_dark"], fill_type="solid")
        area_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
        # Column headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(name="Calibri", size=11, bold=True, color=COLORS["sap_dark"])
            cell.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
            cell.border = Border(
                bottom=Side(style="medium", color=COLORS["sap_dark"]),
                left=Side(style="thin", color=COLORS["border"]),
                right=Side(style="thin", color=COLORS["border"]),
            )
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # Data rows
        for idx, param in enumerate(area_params):
            values = [
                param.get("parameter_name", ""),
                param.get("area", ""),
                param.get("current_value", ""),
                param.get("recommended_value", ""),
                param.get("description", ""),
                param.get("section", ""),
            ]
            
            alt_row = idx % 2 == 1
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=str(value) if value else "")
                cell.font = Font(name="Calibri", size=11)
                cell.border = Border(
                    bottom=Side(style="thin", color=COLORS["border"]),
                    left=Side(style="thin", color=COLORS["border"]),
                    right=Side(style="thin", color=COLORS["border"]),
                )
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if alt_row:
                    cell.fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")
                
                # Highlight recommended value column
                if col == 4 and value:  # Recommended Value column
                    cell.font = Font(name="Calibri", size=11, bold=True, color=COLORS["sap_dark"])
            
            ws.row_dimensions[current_row].height = 35
            current_row += 1
        
        # Blank row between groups
        current_row += 1
    
    # Handle any areas not in the predefined order
    for area, area_params in grouped.items():
        if area in area_order:
            continue
        
        # Area header row
        ws.cell(row=current_row, column=1, value=f"{area} ({len(area_params)} parameters)")
        ws.merge_cells(f"A{current_row}:F{current_row}")
        area_cell = ws[f"A{current_row}"]
        area_cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
        area_cell.fill = PatternFill(start_color=COLORS["sap_dark"], end_color=COLORS["sap_dark"], fill_type="solid")
        current_row += 1
        
        # Column headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        current_row += 1
        
        # Data rows
        for param in area_params:
            values = [
                param.get("parameter_name", ""),
                param.get("area", ""),
                param.get("current_value", ""),
                param.get("recommended_value", ""),
                param.get("description", ""),
                param.get("section", ""),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=str(value) if value else "")
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(wrap_text=True)
            current_row += 1
        
        current_row += 1
    
    # Column widths
    widths = [35, 18, 20, 20, 45, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _normalize_excel_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _looks_like_scalar_value(value: str) -> bool:
    normalized = _normalize_excel_text(value)
    if not normalized:
        return False
    if len(normalized) <= 48 and normalized.count(" ") <= 4:
        return True
    return bool(re.fullmatch(r"[\w\.\-/:=,;]+", normalized)) and len(normalized) <= 96


def _is_narrative_text(value: str) -> bool:
    normalized = _normalize_excel_text(value)
    if not normalized:
        return False
    if len(normalized) >= 80 or normalized.count(" ") >= 8:
        return True
    return any(
        token in f" {normalized.lower()} "
        for token in (" should ", " ensure ", " review ", " update ", " establish ", " perform ", " align ", " monitor ")
    )


def _classify_technical_item(parameter: Dict[str, Any]) -> Dict[str, str]:
    name = _normalize_excel_text(parameter.get("parameter_name"))
    area = _normalize_excel_text(parameter.get("area")) or "General"
    current_value = _normalize_excel_text(parameter.get("current_value"))
    recommended_value = _normalize_excel_text(parameter.get("recommended_value"))
    description = _normalize_excel_text(parameter.get("description"))
    source_section = _normalize_excel_text(parameter.get("source_section") or parameter.get("section"))
    action_status = _normalize_excel_text(parameter.get("action_status")) or "No Action"
    priority = _normalize_excel_text(parameter.get("priority")) or "Low"

    name_lower = name.lower()
    combined_text = f"{description.lower()} {recommended_value.lower()}"

    config_like_name = any(hint in name_lower for hint in CONFIG_NAME_HINTS)
    action_like_name = any(hint in name_lower for hint in ACTION_NAME_HINTS)
    action_like_text = any(hint in combined_text for hint in ACTION_TEXT_HINTS)
    current_is_scalar = _looks_like_scalar_value(current_value)
    target_is_scalar = _looks_like_scalar_value(recommended_value)
    target_is_narrative = _is_narrative_text(recommended_value)

    parameter_score = 0
    action_score = 0

    if config_like_name:
        parameter_score += 3
    if current_is_scalar:
        parameter_score += 1
    if target_is_scalar:
        parameter_score += 1
    if recommended_value and action_status.lower() in {"change required", "verify", "monitor"}:
        parameter_score += 1

    if action_like_name:
        action_score += 3
    if action_like_text:
        action_score += 2
    if target_is_narrative:
        action_score += 2
    if recommended_value and not target_is_scalar:
        action_score += 1
    if any(token in combined_text for token in ("latest", "assigned only", "restore test", "support package", "sp stack")):
        action_score += 2

    if parameter_score >= 4 and action_score <= 2:
        suggested_type = "Parameter"
    elif action_score >= 4 and parameter_score <= 2:
        suggested_type = "Technical Action"
    elif parameter_score >= 3 and action_score >= 3:
        suggested_type = "Mixed"
    elif parameter_score > action_score and parameter_score >= 3:
        suggested_type = "Parameter"
    elif action_score > parameter_score and action_score >= 3:
        suggested_type = "Technical Action"
    else:
        suggested_type = "Needs Review"

    if suggested_type == "Mixed":
        type_confidence = "Low"
        review_flag = "Contains both a settable value and a broader remediation step"
    elif suggested_type == "Needs Review":
        type_confidence = "Low"
        review_flag = "Could not classify confidently"
    else:
        diff = abs(parameter_score - action_score)
        max_score = max(parameter_score, action_score)
        type_confidence = "High" if max_score >= 5 and diff >= 2 else "Medium"
        if suggested_type == "Parameter" and target_is_narrative:
            review_flag = "Review target state wording"
        elif suggested_type == "Technical Action" and config_like_name and target_is_scalar:
            review_flag = "Review whether this belongs on the parameter sheet"
        else:
            review_flag = ""

    recommended_action = description or recommended_value
    if suggested_type in {"Technical Action", "Mixed"} and recommended_value:
        recommended_action = recommended_value
    if not recommended_action:
        recommended_action = name

    return {
        "Original Item": name,
        "Suggested Type": suggested_type,
        "Type Confidence": type_confidence,
        "Area": area,
        "Current State": current_value,
        "Target State": recommended_value,
        "Recommended Action": recommended_action,
        "Description": description,
        "Action Status": action_status,
        "Priority": priority,
        "Source Section": source_section,
        "Review Flag": review_flag,
    }


def _build_technical_register(parameters: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    register: List[Dict[str, str]] = []
    for index, parameter in enumerate(parameters, start=1):
        item = _classify_technical_item(parameter)
        item["Item ID"] = f"TI-{index:02d}"
        register.append(item)
    return register


def _build_technical_register_summary(register: List[Dict[str, str]]) -> str:
    counts = {
        "Parameter": 0,
        "Technical Action": 0,
        "Mixed": 0,
        "Needs Review": 0,
    }
    for item in register:
        suggested_type = item.get("Suggested Type", "Needs Review")
        counts[suggested_type] = counts.get(suggested_type, 0) + 1
    return (
        f"Parameters: {counts.get('Parameter', 0)} | "
        f"Technical Actions: {counts.get('Technical Action', 0)} | "
        f"Mixed: {counts.get('Mixed', 0)} | "
        f"Needs Review: {counts.get('Needs Review', 0)}"
    )


def _tokenize_match_text(value: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9_]+", value.lower())
        if len(token) >= 3 and token not in {
            "the", "and", "for", "with", "that", "this", "from", "into", "only", "than",
            "then", "are", "was", "were", "will", "should", "per", "all", "not", "can",
            "via", "its", "use", "used", "using", "mode", "data", "system", "sap",
        }
    }


def _extract_section_refs(value: str) -> List[str]:
    return re.findall(r"\b\d+(?:\.\d+)+\b", value or "")


def _format_effort_text(value: Any) -> str:
    if isinstance(value, str):
        return _normalize_excel_text(value)
    if isinstance(value, dict):
        return f"Analysis: {value.get('analysis', 'N/A')}, Impl: {value.get('implementation', 'N/A')}"
    return ""


def _build_finding_catalog(data: Dict[str, Any]) -> List[Dict[str, str]]:
    findings = data.get("Key Findings", data.get("key_findings", [])) or []
    recommendations = data.get("Recommendations", data.get("recommendations", [])) or []
    recommendations_by_issue: Dict[str, List[Dict[str, Any]]] = {}
    for recommendation in recommendations:
        linked_issue_id = _normalize_excel_text(
            recommendation.get("Linked issue ID", recommendation.get("linked_issue_id", ""))
        )
        if linked_issue_id:
            recommendations_by_issue.setdefault(linked_issue_id, []).append(recommendation)

    catalog: List[Dict[str, str]] = []
    for finding in findings:
        issue_id = _normalize_excel_text(finding.get("Issue ID", finding.get("issue_id", "")))
        linked_recommendations = recommendations_by_issue.get(issue_id, [])
        primary_recommendation = linked_recommendations[0] if linked_recommendations else {}
        catalog.append(
            {
                "Issue ID": issue_id,
                "Area": _normalize_excel_text(finding.get("Area", finding.get("area", ""))) or "General",
                "Finding": _normalize_excel_text(finding.get("Finding", finding.get("finding", ""))),
                "Impact": _normalize_excel_text(finding.get("Impact", finding.get("impact", ""))),
                "Business Impact": _normalize_excel_text(
                    finding.get("Business impact", finding.get("business_impact", ""))
                ),
                "Severity": _normalize_excel_text(finding.get("Severity", finding.get("severity", "Medium"))) or "Medium",
                "Source": _normalize_excel_text(finding.get("Source", finding.get("source", ""))),
                "Recommendation ID": _normalize_excel_text(
                    primary_recommendation.get("Recommendation ID", primary_recommendation.get("recommendation_id", ""))
                ),
                "Primary Recommendation": _normalize_excel_text(
                    primary_recommendation.get("Action", primary_recommendation.get("action", ""))
                ),
                "Preventative Action": _normalize_excel_text(
                    primary_recommendation.get("Preventative Action", primary_recommendation.get("preventative_action", ""))
                ),
                "Owner Hint": _normalize_excel_text(
                    primary_recommendation.get("Responsible Area", primary_recommendation.get("responsible_area", ""))
                ),
                "Effort": _format_effort_text(
                    primary_recommendation.get("Estimated Effort", primary_recommendation.get("estimated_effort", ""))
                ),
            }
        )
    return catalog


def _looks_like_config_identifier(name: str) -> bool:
    normalized_name = _normalize_excel_text(name)
    if not normalized_name:
        return False
    name_lower = normalized_name.lower()
    if any(hint in name_lower for hint in CONFIG_NAME_HINTS):
        return True
    return bool(re.fullmatch(r"[A-Z0-9_./\[\]-]+", normalized_name))


def _infer_action_type(item: Dict[str, str]) -> str:
    name = _normalize_excel_text(item.get("Original Item"))
    description = _normalize_excel_text(item.get("Description"))
    source_section = _normalize_excel_text(item.get("Source Section"))
    suggested_type = _normalize_excel_text(item.get("Suggested Type"))
    text = f"{name} {description} {source_section}".lower()
    if "job" in text or "schedule" in text or "/ui" in name.lower() or "/uif" in name.lower():
        return "Scheduled Job"
    if suggested_type == "Parameter" or _looks_like_config_identifier(name):
        return "Config Parameter"
    if any(token in text for token in ("privilege", "authorization", "password", "user", "role", "acl", "gateway security")):
        return "Authorization / Access"
    if any(token in text for token in ("support package", "sapui5", "kernel", "release", "patch", "upgrade", "version", "windows server")):
        return "Patch / Upgrade"
    if any(token in text for token in ("cpu", "memory", "ping", "response time", "monitor", "threshold", "latency", "workload", "capacity")):
        return "Capacity / Monitoring"
    if any(token in text for token in ("ssl", "tls", "network", "listeninterface", "icm/", "communication", "hostname")):
        return "Network / Infrastructure"
    return "Process / Governance"


def _get_action_display_text(item: Dict[str, str]) -> str:
    item_id = _normalize_excel_text(item.get("Item ID"))
    original_item = _normalize_excel_text(item.get("Original Item"))
    recommended_action = _normalize_excel_text(item.get("Recommended Action"))
    target_state = _normalize_excel_text(item.get("Target State"))
    if item_id.startswith("REC-") or item_id.startswith("KF-"):
        return recommended_action or original_item
    return original_item or recommended_action or target_state


def _match_technical_item_to_finding(item: Dict[str, str], finding_catalog: List[Dict[str, str]]) -> Dict[str, str]:
    item_text = " ".join(
        [
            item.get("Original Item", ""),
            item.get("Description", ""),
            item.get("Recommended Action", ""),
            item.get("Source Section", ""),
            item.get("Area", ""),
        ]
    ).lower()
    item_tokens = _tokenize_match_text(item_text)
    item_sections = set(_extract_section_refs(item.get("Source Section", "")))

    best_match: Dict[str, str] = {}
    best_score = 0
    for finding in finding_catalog:
        score = 0
        finding_sections = set(_extract_section_refs(finding.get("Source", "")))
        if item_sections and finding_sections:
            shared_sections = item_sections & finding_sections
            if shared_sections:
                score += 8 + len(shared_sections)
            elif any(
                item_section.startswith(finding_section) or finding_section.startswith(item_section)
                for item_section in item_sections
                for finding_section in finding_sections
            ):
                score += 5
        if item.get("Area", "").lower() == finding.get("Area", "").lower():
            score += 2
        finding_tokens = _tokenize_match_text(
            f"{finding.get('Finding', '')} {finding.get('Source', '')} {finding.get('Primary Recommendation', '')}"
        )
        score += min(len(item_tokens & finding_tokens), 4)
        if finding.get("Finding", "").lower() and finding.get("Finding", "").lower() in item_text:
            score += 3
        if score > best_score:
            best_score = score
            best_match = finding

    return best_match if best_score >= 4 else {}


def _build_enriched_technical_items(
    technical_register: List[Dict[str, str]],
    finding_catalog: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    enriched_items: List[Dict[str, str]] = []
    for item in technical_register:
        finding_match = _match_technical_item_to_finding(item, finding_catalog)
        enriched_item = dict(item)
        enriched_item["Action Type"] = _infer_action_type(enriched_item)
        enriched_item["Finding ID"] = finding_match.get("Issue ID", "")
        enriched_item["Finding"] = finding_match.get("Finding", "")
        enriched_item["Owner Hint"] = finding_match.get("Owner Hint", "")
        enriched_item["Effort"] = finding_match.get("Effort", "")
        enriched_item["Primary Recommendation"] = finding_match.get("Primary Recommendation", "")
        enriched_items.append(enriched_item)
    return enriched_items


def _is_reference_note(item: Dict[str, str]) -> bool:
    status = _normalize_excel_text(item.get("Action Status", "")).lower()
    combined_text = " ".join(
        [
            item.get("Original Item", ""),
            item.get("Description", ""),
            item.get("Recommended Action", ""),
            item.get("Target State", ""),
            item.get("Source Section", ""),
        ]
    ).lower()
    if status == "no action":
        return True
    if any(token in combined_text for token in ("threshold", "reference", "benchmark", "guideline", "response time", "ping", "latency")):
        return status in {"monitor", "no action"}
    return False


def _build_fallback_action_item(finding: Dict[str, str]) -> Dict[str, str]:
    fallback_action = finding.get("Primary Recommendation") or finding.get("Finding", "")
    fallback_item = {
        "Item ID": finding.get("Recommendation ID") or finding.get("Issue ID", ""),
        "Original Item": finding.get("Finding", ""),
        "Suggested Type": "Technical Action",
        "Type Confidence": "High",
        "Area": finding.get("Area", "General"),
        "Current State": "",
        "Target State": "",
        "Recommended Action": fallback_action,
        "Description": finding.get("Impact", ""),
        "Action Status": "Change Required",
        "Priority": finding.get("Severity", "Medium").title(),
        "Source Section": finding.get("Source", ""),
        "Review Flag": "",
        "Finding ID": finding.get("Issue ID", ""),
        "Finding": finding.get("Finding", ""),
        "Owner Hint": finding.get("Owner Hint", ""),
        "Effort": finding.get("Effort", ""),
        "Primary Recommendation": fallback_action,
    }
    fallback_item["Action Type"] = _infer_action_type(fallback_item)
    return fallback_item


def _build_action_register_items(
    enriched_items: List[Dict[str, str]],
    finding_catalog: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    action_items = [item for item in enriched_items if not _is_reference_note(item)]
    covered_findings = {item.get("Finding ID", "") for item in action_items if item.get("Finding ID")}
    for finding in finding_catalog:
        if finding.get("Issue ID") and finding.get("Issue ID") not in covered_findings:
            action_items.append(_build_fallback_action_item(finding))

    finding_order = {finding.get("Issue ID", ""): index for index, finding in enumerate(finding_catalog)}
    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    action_items.sort(
        key=lambda item: (
            finding_order.get(item.get("Finding ID", ""), len(finding_order) + 1),
            priority_order.get(item.get("Priority", "Low"), 3),
            item.get("Action Type", ""),
            item.get("Item ID", ""),
        )
    )
    return action_items


def _build_reference_note_items(enriched_items: List[Dict[str, str]]) -> List[Dict[str, str]]:
    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    reference_items = [item for item in enriched_items if _is_reference_note(item)]
    reference_items.sort(
        key=lambda item: (
            item.get("Finding ID", "ZZZ"),
            priority_order.get(item.get("Priority", "Low"), 3),
            item.get("Action Type", ""),
            item.get("Item ID", ""),
        )
    )
    return reference_items


def _write_actions_by_finding_sheet(
    ws: Worksheet,
    finding_catalog: List[Dict[str, str]],
    action_items: List[Dict[str, str]],
    styles: Dict[str, NamedStyle],
) -> Dict[str, int]:
    ws.title = "Technical Actions by Finding"
    ws["A1"] = "Technical Actions by Finding"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 25

    if not finding_catalog and not action_items:
        ws["A3"] = "No technical actions found."
        return {}

    headers = [
        "Item ID",
        "Action Type",
        "Technical Action",
        "Current State",
        "Target State",
        "Priority",
        "Owner Hint",
        "Source Section",
        "Status",
    ]
    actions_by_finding: Dict[str, List[Dict[str, str]]] = {}
    for item in action_items:
        actions_by_finding.setdefault(item.get("Finding ID", ""), []).append(item)

    current_row = 3
    anchors: Dict[str, int] = {}
    for finding in finding_catalog:
        finding_id = finding.get("Issue ID", "")
        linked_actions = actions_by_finding.get(finding_id, []) or [_build_fallback_action_item(finding)]
        anchors[finding_id] = current_row
        ws.cell(row=current_row, column=1, value=f"{finding_id} | {finding.get('Area', '')} | {finding.get('Finding', '')}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
        title_cell.fill = _get_severity_fill(finding.get("Severity", "Medium"))
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        summary_parts = [
            f"Impact: {finding.get('Impact', '')}",
            f"Business impact: {finding.get('Business Impact', '')}",
        ]
        if finding.get("Primary Recommendation"):
            summary_parts.append(f"Primary recommendation: {finding.get('Primary Recommendation', '')}")
        if finding.get("Owner Hint"):
            summary_parts.append(f"Owner hint: {finding.get('Owner Hint', '')}")
        if finding.get("Effort"):
            summary_parts.append(f"Effort: {finding.get('Effort', '')}")
        ws.cell(row=current_row, column=1, value="\n".join(part for part in summary_parts if part and not part.endswith(": ")))
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        summary_cell = ws.cell(row=current_row, column=1)
        summary_cell.font = Font(name="Calibri", size=11, italic=True, color=COLORS["sap_dark"])
        summary_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[current_row].height = 78
        current_row += 1

        for col_index, header in enumerate(headers, start=1):
            header_cell = ws.cell(row=current_row, column=col_index, value=header)
            _apply_header_style(header_cell, styles)
        current_row += 1

        action_start_row = current_row
        for action_index, item in enumerate(linked_actions):
            values = [
                item.get("Item ID", ""),
                item.get("Action Type", ""),
                _get_action_display_text(item),
                item.get("Current State", ""),
                item.get("Target State", ""),
                item.get("Priority", ""),
                item.get("Owner Hint", "") or finding.get("Owner Hint", ""),
                item.get("Source Section", "") or finding.get("Source", ""),
                item.get("Action Status", ""),
            ]
            alt_row = action_index % 2 == 1
            for col_index, value in enumerate(values, start=1):
                data_cell = ws.cell(row=current_row, column=col_index, value=str(value) if value else "")
                _apply_data_style(data_cell, styles, alt_row)
                if col_index == 3:
                    data_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            ws.row_dimensions[current_row].height = 42
            current_row += 1

        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.row_dimensions.group(action_start_row, current_row - 1, outline_level=1, hidden=False)
        current_row += 1

    unlinked_actions = [item for item in action_items if not item.get("Finding ID")]
    if unlinked_actions:
        ws.cell(row=current_row, column=1, value="Additional technical actions")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        extra_title_cell = ws.cell(row=current_row, column=1)
        extra_title_cell.font = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
        extra_title_cell.fill = PatternFill(start_color=COLORS["sap_dark"], end_color=COLORS["sap_dark"], fill_type="solid")
        current_row += 1
        for col_index, header in enumerate(headers, start=1):
            header_cell = ws.cell(row=current_row, column=col_index, value=header)
            _apply_header_style(header_cell, styles)
        current_row += 1
        for action_index, item in enumerate(unlinked_actions):
            values = [
                item.get("Item ID", ""),
                item.get("Action Type", ""),
                _get_action_display_text(item),
                item.get("Current State", ""),
                item.get("Target State", ""),
                item.get("Priority", ""),
                item.get("Owner Hint", ""),
                item.get("Source Section", ""),
                item.get("Action Status", ""),
            ]
            alt_row = action_index % 2 == 1
            for col_index, value in enumerate(values, start=1):
                data_cell = ws.cell(row=current_row, column=col_index, value=str(value) if value else "")
                _apply_data_style(data_cell, styles, alt_row)
            ws.row_dimensions[current_row].height = 42
            current_row += 1

    for col_index, width in enumerate([12, 20, 44, 24, 24, 12, 24, 28, 14], start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.freeze_panes = "A4"
    return anchors


def _write_tabular_sheet(
    ws: Worksheet,
    title: str,
    headers: List[str],
    rows: List[List[str]],
    styles: Dict[str, NamedStyle],
    widths: List[int],
    empty_message: str,
    subtitle: str = "",
):
    ws.title = title
    last_col_letter = get_column_letter(len(headers))

    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.row_dimensions[1].height = 25

    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Calibri", size=11, italic=True, color=COLORS["sap_dark"])
        ws.merge_cells(f"A2:{last_col_letter}2")

    header_row = 3
    if not rows:
        ws[f"A{header_row}"] = empty_message
        ws.merge_cells(f"A{header_row}:{last_col_letter}{header_row}")
        return

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        _apply_header_style(cell, styles)

    current_row = header_row + 1
    for row_index, row_values in enumerate(rows):
        alt_row = row_index % 2 == 1
        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_index, value=str(value) if value else "")
            _apply_data_style(cell, styles, alt_row)
        ws.row_dimensions[current_row].height = 34
        current_row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{current_row - 1}"
    for col_index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width


def json_to_excel(
    json_data: Dict[str, Any],
    customer_name: str = "",
    parameters: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """
    Convert EWA analysis JSON to a formatted Excel workbook.
    
    Args:
        json_data: The EWA analysis JSON data
        customer_name: Optional customer name for the report
        parameters: Optional list of parameter recommendations extracted from markdown
        
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    styles = _create_styles()
    
    # Register styles with workbook
    for style in styles.values():
        try:
            wb.add_named_style(style)
        except ValueError:
            pass  # Style already exists
    
    # Create sheets
    # Remove default sheet and create our own
    default_sheet = wb.active
    
    # 1. Summary sheet
    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, json_data, styles, customer_name)
    
    # 2. Positive Findings
    ws_positive = wb.create_sheet("Positive Findings")
    _write_positive_findings_sheet(ws_positive, json_data, styles)
    
    # 3. Findings & Actions (combined Key Findings + Recommendations)
    ws_findings = wb.create_sheet("Findings & Actions")
    _write_findings_and_recommendations_sheet(ws_findings, json_data, styles)
    
    technical_register = _build_technical_register(parameters or [])
    finding_catalog = _build_finding_catalog(json_data)
    enriched_technical_items = _build_enriched_technical_items(technical_register, finding_catalog)
    action_register_items = _build_action_register_items(enriched_technical_items, finding_catalog)
    reference_note_items = _build_reference_note_items(enriched_technical_items)

    # 4. Technical register and filtered technical views (if available)
    if finding_catalog or action_register_items or reference_note_items:
        ws_register = wb.create_sheet("Technical Actions by Finding")
        finding_anchors = _write_actions_by_finding_sheet(
            ws_register,
            finding_catalog,
            action_register_items,
            styles,
        )

        ws_params = wb.create_sheet("Action Register")
        _write_tabular_sheet(
            ws_params,
            "Action Register",
            [
                "Finding ID",
                "Area",
                "Finding",
                "Action Type",
                "Technical Action",
                "Current State",
                "Target State",
                "Priority",
                "Owner Hint",
                "Source Section",
                "Status",
            ],
            [[item.get(key, "") for key in [
                "Finding ID",
                "Area",
                "Finding",
                "Action Type",
            ]] + [
                _get_action_display_text(item),
                item.get("Current State", ""),
                item.get("Target State", ""),
                item.get("Priority", ""),
                item.get("Owner Hint", ""),
                item.get("Source Section", ""),
                item.get("Action Status", ""),
            ] for item in action_register_items],
            styles,
            [12, 18, 28, 20, 40, 22, 22, 12, 24, 28, 14],
            "No action items found.",
        )

        for row_index, item in enumerate(action_register_items, start=4):
            finding_id = item.get("Finding ID", "")
            if finding_id and finding_id in finding_anchors:
                link_cell = ws_params.cell(row=row_index, column=1)
                link_cell.hyperlink = f"#'Technical Actions by Finding'!A{finding_anchors[finding_id]}"
                link_cell.style = "Hyperlink"

        ws_review = wb.create_sheet("Reference Notes")
        _write_tabular_sheet(
            ws_review,
            "Reference Notes",
            [
                "Finding ID",
                "Area",
                "Finding",
                "Note Type",
                "Technical Note",
                "Current State",
                "Reference Value",
                "Source Section",
                "Status",
                "Review Flag",
            ],
            [[item.get(key, "") for key in [
                "Finding ID",
                "Area",
                "Finding",
                "Action Type",
                "Recommended Action",
                "Current State",
                "Target State",
                "Source Section",
                "Action Status",
                "Review Flag",
            ]] for item in reference_note_items],
            styles,
            [12, 18, 28, 20, 40, 22, 22, 28, 14, 28],
            "No reference notes found.",
        )
    
    # 5. Capacity Outlook
    ws_capacity = wb.create_sheet("Capacity Outlook")
    _write_capacity_sheet(ws_capacity, json_data, styles)
    
    # 6. Chapters Reviewed
    ws_chapters = wb.create_sheet("Chapters Reviewed")
    _write_chapters_sheet(ws_chapters, json_data, styles)
    
    # Remove the default sheet
    wb.remove(default_sheet)
    
    # Set Summary as active sheet
    wb.active = wb["Summary"]
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
