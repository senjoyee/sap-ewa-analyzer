"""
8-Tab Excel workbook generator for v2.0 pillar-based EWA analysis.

Tab layout:
  1. Executive Summary — metadata, health overview, top-3 risks, audit trail
  2. Security & Compliance
  3. Basis Core Operations
  4. Database & Infrastructure
  5. Integration & Connectivity
  6. Lifecycle & Upgrades
  7. Business Processes & DVM
  8. Uncategorized

Tabs 2-8 share the same structure: Findings table, Positives table,
Recommendations table, and an optional Parameter Changes table.

Reuses the SAP color palette and helper functions from ``excel_utils``.
"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.excel_utils import (
    COLORS,
    _create_styles,
    _apply_header_style,
    _apply_data_style,
    _get_severity_fill,
    _get_status_fill,
)
from models.pillar_models import (
    PillarEnum,
    PILLAR_DISPLAY_NAMES,
    PILLAR_TAB_ICONS,
    PILLAR_PERSONAS,
    PILLAR_ORDER,
)


# Pillar tabs in workbook order (executive_summary is separate Tab 1)
_PILLAR_TAB_ORDER: list[str] = [
    PillarEnum.SECURITY_COMPLIANCE,
    PillarEnum.BASIS_OPERATIONS,
    PillarEnum.DATABASE_INFRASTRUCTURE,
    PillarEnum.INTEGRATION_CONNECTIVITY,
    PillarEnum.LIFECYCLE_UPGRADES,
    PillarEnum.BUSINESS_PROCESSES_DVM,
    PillarEnum.UNCATEGORIZED,
]


# ─── Tab 1: Executive Summary ────────────────────────────────────────────

def _write_executive_tab(
    ws: Worksheet,
    data: Dict[str, Any],
    styles: Dict[str, NamedStyle],
    customer_name: str = "",
) -> None:
    ws.title = f"{PILLAR_TAB_ICONS[PillarEnum.EXECUTIVE_SUMMARY]} Executive Summary"

    # ---- Title ----
    ws["A1"] = "EWA Analysis — Executive Summary"
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32

    # ---- System Metadata ----
    ws["A3"] = "System Information"
    ws["A3"].font = Font(name="Calibri", size=14, bold=True, color=COLORS["sap_dark"])
    ws.merge_cells("A3:E3")

    meta = data.get("System Metadata", {}) or {}
    row = 4

    if customer_name:
        _kv(ws, row, "Customer", customer_name, styles)
        row += 1

    for label, key in [("System ID", "System ID"), ("Report Date", "Report Date"), ("Analysis Period", "Analysis Period")]:
        _kv(ws, row, label, meta.get(key, "N/A"), styles)
        row += 1

    # Overall Risk
    row += 1
    overall_risk = str(data.get("Overall Risk", "N/A"))
    _kv(ws, row, "Overall Risk", overall_risk.upper(), styles)
    risk_cell = ws.cell(row=row, column=2)
    risk_cell.font = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
    risk_cell.fill = _get_severity_fill(overall_risk)
    risk_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # ---- System Health Overview ----
    ws.cell(row=row, column=1, value="System Health Overview").font = Font(
        name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
    )
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    health = data.get("System Health Overview", {}) or {}
    for dim in ["Performance", "Security", "Stability", "configuration"]:
        status = health.get(dim, health.get(dim.lower()))
        if status:
            ws.cell(row=row, column=1, value=dim.title()).font = styles["label"].font
            c = ws.cell(row=row, column=2, value=str(status).upper())
            c.font = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
            c.fill = _get_status_fill(str(status))
            c.alignment = Alignment(horizontal="center", vertical="center")
            row += 1

    row += 1

    # ---- Executive Summary text ----
    ws.cell(row=row, column=1, value="Executive Summary").font = Font(
        name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
    )
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    exec_text = data.get("Executive Summary", "")
    if exec_text:
        c = ws.cell(row=row, column=1, value=exec_text)
        c.font = styles["value"].font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = max(60, len(exec_text) // 3)
        row += 2

    # ---- Top 3 Critical Risks ----
    risks = data.get("Top 3 Critical Risks", []) or []
    if risks:
        ws.cell(row=row, column=1, value="Top 3 Critical Risks").font = Font(
            name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
        )
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        risk_headers = ["#", "Risk", "Pillar", "Severity"]
        for col, h in enumerate(risk_headers, 1):
            _apply_header_style(ws.cell(row=row, column=col, value=h), styles)
        row += 1

        for idx, risk in enumerate(risks, 1):
            pillar_display = PILLAR_DISPLAY_NAMES.get(risk.get("pillar", ""), risk.get("pillar", ""))
            values = [str(idx), risk.get("risk", ""), pillar_display, risk.get("severity", "")]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                _apply_data_style(cell, styles, idx % 2 == 0)
                if col == 4:
                    cell.font = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
                    cell.fill = _get_severity_fill(v)
                    cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[row].height = 40
            row += 1
        row += 1

    # ---- Audit Trail ----
    audit = data.get("Audit Trail", {}) or {}
    if audit:
        ws.cell(row=row, column=1, value="Audit Trail").font = Font(
            name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
        )
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        _kv(ws, row, "RED/YELLOW Total", str(audit.get("red_yellow_total", "?")), styles); row += 1
        _kv(ws, row, "RED/YELLOW Mapped", str(audit.get("red_yellow_mapped", "?")), styles); row += 1
        _kv(ws, row, "Coverage %", f"{audit.get('coverage_pct', '?')}%", styles); row += 1

        unmapped = audit.get("unmapped_chapters", [])
        if unmapped:
            _kv(ws, row, "Unmapped Chapters", ", ".join(unmapped), styles)
            row += 1
        row += 1

    # ---- Capacity Outlook ----
    capacity = data.get("Capacity Outlook", {}) or {}
    if capacity:
        ws.cell(row=row, column=1, value="Capacity Outlook").font = Font(
            name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
        )
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        for label, key in [
            ("Database Growth", "Database Growth"),
            ("CPU Utilization", "CPU Utilization"),
            ("Memory Utilization", "Memory Utilization"),
            ("Summary", "Summary"),
        ]:
            val = capacity.get(key, "")
            if val:
                _kv(ws, row, label, str(val), styles)
                row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15


# ─── Tabs 2-8: Pillar sheets ─────────────────────────────────────────────

def _write_pillar_tab(
    ws: Worksheet,
    pillar_key: str,
    data: Dict[str, Any],
    styles: Dict[str, NamedStyle],
) -> None:
    display = PILLAR_DISPLAY_NAMES.get(pillar_key, pillar_key)
    icon = PILLAR_TAB_ICONS.get(pillar_key, "")
    persona = PILLAR_PERSONAS.get(pillar_key, "")
    ws.title = f"{icon} {display}"

    # ---- Title ----
    ws["A1"] = display
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=COLORS["sap_gold"])
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Target: {persona}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells("A2:I2")

    pillar_bucket = (data.get("Pillars") or {}).get(pillar_key, {})
    findings = pillar_bucket.get("findings", []) or []
    positives = pillar_bucket.get("positives", []) or []
    recs = pillar_bucket.get("recommendations", []) or []
    params = pillar_bucket.get("parameters", []) or []

    row = 4

    # ──── Findings table ────
    row = _write_findings_section(ws, findings, recs, row, styles)

    # ──── Positive Findings ────
    if positives:
        row += 1
        row = _write_positives_section(ws, positives, row, styles)

    # ──── Parameter Changes ────
    if params:
        row += 1
        row = _write_params_section(ws, params, row, styles)

    # Column widths — standard for all pillar tabs
    widths = [10, 18, 35, 28, 28, 20, 35, 35, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_findings_section(
    ws: Worksheet,
    findings: List[Dict[str, Any]],
    recs: List[Dict[str, Any]],
    start_row: int,
    styles: Dict[str, NamedStyle],
) -> int:
    """Write findings grouped by priority with linked recommendations."""
    row = start_row

    ws.cell(row=row, column=1, value="Findings & Recommended Actions").font = Font(
        name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
    )
    ws.merge_cells(f"A{row}:I{row}")
    row += 1

    if not findings:
        ws.cell(row=row, column=1, value="No findings in this pillar.").font = styles["value"].font
        return row + 1

    # Build rec lookup by Linked issue ID
    recs_by_issue: Dict[str, Dict] = {}
    for rec in recs:
        lid = rec.get("Linked issue ID", "")
        if lid and lid not in recs_by_issue:
            recs_by_issue[lid] = rec

    # Group by priority
    priority_order = ["CRITICAL", "HIGH", "LOW/INFO"]
    grouped: Dict[str, list] = {p: [] for p in priority_order}
    for f in findings:
        p = f.get("priority", "LOW/INFO")
        if p not in grouped:
            p = "LOW/INFO"
        grouped[p].append(f)

    headers = [
        "Issue ID", "Category", "Finding", "Impact", "Business Impact",
        "Reference", "Action", "Preventative Action", "Effort",
    ]

    for priority in priority_order:
        group = grouped[priority]
        if not group:
            continue

        # Priority banner
        severity_for_color = {"CRITICAL": "critical", "HIGH": "high", "LOW/INFO": "medium"}.get(priority, "medium")
        banner = ws.cell(row=row, column=1, value=f"{priority}  ({len(group)})")
        ws.merge_cells(f"A{row}:I{row}")
        banner.font = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
        banner.fill = _get_severity_fill(severity_for_color)
        banner.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        # Column headers
        for col, h in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=row, column=col, value=h), styles)
        row += 1

        # Data rows
        for idx, finding in enumerate(group):
            issue_id = finding.get("Issue ID", "")
            rec = recs_by_issue.get(issue_id, {})
            effort = rec.get("Estimated Effort", {})
            if isinstance(effort, dict):
                effort_str = f"A: {effort.get('analysis','?')}, I: {effort.get('implementation','?')}"
            else:
                effort_str = str(effort) if effort else ""

            values = [
                issue_id,
                finding.get("category", ""),
                finding.get("Finding", ""),
                finding.get("Impact", ""),
                finding.get("Business impact", ""),
                finding.get("reference", ""),
                rec.get("Action", ""),
                rec.get("Preventative Action", ""),
                effort_str,
            ]

            alt = idx % 2 == 1
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=str(v) if v else "")
                _apply_data_style(cell, styles, alt)

            ws.row_dimensions[row].height = 60
            row += 1
        row += 1  # blank between groups

    return row


def _write_positives_section(
    ws: Worksheet,
    positives: List[Dict[str, Any]],
    start_row: int,
    styles: Dict[str, NamedStyle],
) -> int:
    row = start_row
    ws.cell(row=row, column=1, value="Positive Findings").font = Font(
        name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
    )
    ws.merge_cells(f"A{row}:I{row}")
    row += 1

    headers = ["Area", "Description"]
    for col, h in enumerate(headers, 1):
        _apply_header_style(ws.cell(row=row, column=col, value=h), styles)
    row += 1

    for idx, pos in enumerate(positives):
        ws.cell(row=row, column=1, value=pos.get("Area", "")).font = styles["label"].font
        c = ws.cell(row=row, column=2, value=pos.get("Description", ""))
        _apply_data_style(c, styles, idx % 2 == 1)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 35
        row += 1

    return row


def _write_params_section(
    ws: Worksheet,
    params: List[Dict[str, Any]],
    start_row: int,
    styles: Dict[str, NamedStyle],
) -> int:
    row = start_row
    ws.cell(row=row, column=1, value="Parameter Changes").font = Font(
        name="Calibri", size=14, bold=True, color=COLORS["sap_dark"]
    )
    ws.merge_cells(f"A{row}:I{row}")
    row += 1

    headers = ["Parameter", "Area", "Current", "Recommended", "Description", "Section"]
    for col, h in enumerate(headers, 1):
        _apply_header_style(ws.cell(row=row, column=col, value=h), styles)
    row += 1

    for idx, p in enumerate(params):
        values = [
            p.get("parameter_name", ""),
            p.get("area", ""),
            p.get("current_value", ""),
            p.get("recommended_value", ""),
            p.get("description", ""),
            p.get("section", ""),
        ]
        alt = idx % 2 == 1
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(v) if v else "")
            _apply_data_style(cell, styles, alt)
            if col == 4 and v:
                cell.font = Font(name="Calibri", size=11, bold=True, color=COLORS["sap_dark"])
        ws.row_dimensions[row].height = 30
        row += 1

    return row


# ─── Helper ──────────────────────────────────────────────────────────────

def _kv(ws: Worksheet, row: int, label: str, value: str, styles: Dict[str, NamedStyle]) -> None:
    """Write a key-value pair in columns A-B."""
    ws.cell(row=row, column=1, value=label).font = styles["label"].font
    ws.cell(row=row, column=2, value=value).font = styles["value"].font


# ─── Public API ──────────────────────────────────────────────────────────

def json_to_pillar_excel(
    json_data: Dict[str, Any],
    customer_name: str = "",
    parameters: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """Convert v2.0 pillar-based EWA JSON to a formatted 8-tab Excel workbook.

    Args:
        json_data: The v2.0 EWA analysis JSON data.
        customer_name: Optional customer name for the title page.
        parameters: Optional parameter extractions (already distributed into pillars if
                    ``distribute_parameters()`` was called earlier).

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()
    styles = _create_styles()

    for style in styles.values():
        try:
            wb.add_named_style(style)
        except ValueError:
            pass

    default_sheet = wb.active

    # Tab 1 — Executive Summary
    ws_exec = wb.create_sheet()
    _write_executive_tab(ws_exec, json_data, styles, customer_name)

    # Tabs 2-8 — Pillar sheets
    for pillar_key in _PILLAR_TAB_ORDER:
        ws = wb.create_sheet()
        _write_pillar_tab(ws, pillar_key, json_data, styles)

    # Remove default blank sheet
    wb.remove(default_sheet)

    # Set Executive Summary as active
    wb.active = 0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
